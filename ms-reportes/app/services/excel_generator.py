from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_calificaciones_excel(materia_data: dict, alumnos: list, concentrado_alumnos: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calificaciones"
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Reporte de Calificaciones"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Materia info
    ws['A3'] = "Materia:"
    ws['B3'] = materia_data.get('nombre', 'N/A')
    ws['A4'] = "NRC:"
    ws['B4'] = materia_data.get('nrc', 'N/A')
    ws['A5'] = "Docente:"
    ws['B5'] = materia_data.get('docente_nombre', 'N/A')
    
    if materia_data.get('periodo'):
        ws['A6'] = "Periodo:"
        ws['B6'] = materia_data['periodo'].get('nombre', 'N/A')
    
    # Headers
    headers = ["Matrícula", "Nombre Alumno", "Promedio Real", "Promedio Final", "Estado"]
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    concentrado_dict = {
        a.alumno_id: {
            'real': a.promedio_real,
            'redondeado': a.promedio_redondeado
        } for a in concentrado_alumnos
    }
    
    row_num = 9
    for al in alumnos:
        c_data = concentrado_dict.get(al.alumno_id, {})
        prom_real = c_data.get('real', 0.0)
        prom_redon = c_data.get('redondeado', 0)
        estado = "Aprobado" if prom_redon >= 6 else "Reprobado"
        
        ws.cell(row=row_num, column=1, value=al.matricula)
        ws.cell(row=row_num, column=2, value=al.nombre_completo)
        ws.cell(row=row_num, column=3, value=prom_real)
        ws.cell(row=row_num, column=4, value=prom_redon)
        ws.cell(row=row_num, column=5, value=estado)
        row_num += 1
        
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 6):
        column = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(7, row_num):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                if len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
        ws.column_dimensions[column].width = max_length + 2

    buffer = BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes

def generate_asistencias_excel(materia_data: dict, alumnos: list, asistencias_data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"
    
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Reporte de Asistencias"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    ws['A3'] = "Materia:"
    ws['B3'] = materia_data.get('nombre', 'N/A')
    ws['A4'] = "NRC:"
    ws['B4'] = materia_data.get('nrc', 'N/A')
    ws['A5'] = "Docente:"
    ws['B5'] = materia_data.get('docente_nombre', 'N/A')
    
    if materia_data.get('periodo'):
        ws['A6'] = "Periodo:"
        ws['B6'] = materia_data['periodo'].get('nombre', 'N/A')
    
    headers = ["Matrícula", "Nombre Alumno", "Presentes", "Retardos", "Faltas", "% Asistencia"]
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    row_num = 9
    for al in alumnos:
        a_data = asistencias_data.get(al.alumno_id, {})
        presentes = a_data.get("presentes", 0)
        retardos = a_data.get("retardos", 0)
        faltas = a_data.get("faltas", 0)
        porcentaje = a_data.get("porcentaje", 0.0)
        
        ws.cell(row=row_num, column=1, value=al.matricula)
        ws.cell(row=row_num, column=2, value=al.nombre_completo)
        ws.cell(row=row_num, column=3, value=presentes)
        ws.cell(row=row_num, column=4, value=retardos)
        ws.cell(row=row_num, column=5, value=faltas)
        ws.cell(row=row_num, column=6, value=f"{porcentaje:.1f}%")
        row_num += 1
        
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 7):
        column = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(7, row_num):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                if len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
        ws.column_dimensions[column].width = max_length + 2

    buffer = BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes
